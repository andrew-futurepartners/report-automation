"""Tests for crosstab_parser.py — Excel workbook parsing into table dicts."""

import json

import pytest
from openpyxl import Workbook

from crosstab_parser import parse_workbook, to_json


def _save_wb(wb, tmp_path, name="test.xlsx"):
    path = tmp_path / name
    wb.save(str(path))
    return str(path)


def _make_standard_wb():
    """Workbook with a title row, metric row, banner row, and body rows in a single block.

    The banner row has more non-null cells than data rows (which have some
    None values) so the parser's max-non-null heuristic correctly identifies it.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Row 1: title (single cell in col A → title-in-first-row pattern)
    ws.cell(row=1, column=1, value="Q1 Brand Awareness")

    # Row 2: metric row (group labels, sparse — 3 non-nulls)
    ws.cell(row=2, column=1, value="Metric")
    ws.cell(row=2, column=2, value="Gender")
    ws.cell(row=2, column=5, value="Age")

    # Row 3: banner row (8 non-nulls, label cell + 7 column headers)
    ws.cell(row=3, column=1, value="Segment")
    ws.cell(row=3, column=2, value="Male")
    ws.cell(row=3, column=3, value="Female")
    ws.cell(row=3, column=4, value="Other")
    ws.cell(row=3, column=5, value="18-34")
    ws.cell(row=3, column=6, value="35-54")
    ws.cell(row=3, column=7, value="55+")
    ws.cell(row=3, column=8, value="Total")

    # Rows 4-6: body (label + 5 data values + 2 None gaps = 6 non-nulls each)
    for r, (label, *vals) in enumerate([
        ("Aided",       80, 85, None, 75, 82, None, 83),
        ("Unaided",     40, 45, None, 38, 42, None, 41),
        ("Top of Mind", 20, 22, None, 18, 21, None, 21),
    ], start=4):
        ws.cell(row=r, column=1, value=label)
        for c, v in enumerate(vals, start=2):
            if v is not None:
                ws.cell(row=r, column=c, value=v)

    return wb


# ---------------------------------------------------------------------------
# Standard two-row header
# ---------------------------------------------------------------------------

class TestStandardTwoRowHeader:
    def test_parses_title_and_labels(self, tmp_path):
        wb = _make_standard_wb()
        path = _save_wb(wb, tmp_path)
        result = parse_workbook(path)
        tables = result["tables"]
        assert len(tables) >= 1
        t = tables[0]
        assert t["title"] == "Q1 Brand Awareness"
        assert "Aided" in t["row_labels"]
        assert "Unaided" in t["row_labels"]
        assert "Top of Mind" in t["row_labels"]
        assert len(t["col_labels"]) >= 7
        # Two-row header: col labels include pipe-separated group names
        assert any("|" in cl for cl in t["col_labels"])

    def test_values_are_numeric(self, tmp_path):
        wb = _make_standard_wb()
        path = _save_wb(wb, tmp_path)
        t = parse_workbook(path)["tables"][0]
        for row in t["values"]:
            for val in row:
                assert val is None or isinstance(val, (int, float))


# ---------------------------------------------------------------------------
# Single-row header (no metric row)
# ---------------------------------------------------------------------------

class TestSingleRowHeader:
    def test_col_labels_from_banner_only(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Title row (single cell → consumed as title)
        ws.cell(row=1, column=1, value="Simple Table")
        # Banner row: dense (first row of remaining block after title consumed)
        ws.cell(row=2, column=1, value="")
        ws.cell(row=2, column=2, value="Col A")
        ws.cell(row=2, column=3, value="Col B")
        ws.cell(row=2, column=4, value="Col C")

        # Body rows (more non-null per row than banner has, so banner stays banner)
        for r, (label, *vals) in enumerate([
            ("Row 1", 10, 20, 30),
            ("Row 2", 40, 50, 60),
            ("Row 3", 70, 80, 90),
        ], start=3):
            ws.cell(row=r, column=1, value=label)
            for c, v in enumerate(vals, start=2):
                ws.cell(row=r, column=c, value=v)

        path = _save_wb(wb, tmp_path)
        t = parse_workbook(path)["tables"][0]
        # The banner row has "Col A", "Col B", "Col C"; no metric row → no pipe separator
        found_labels = t["col_labels"]
        assert any("Col A" in cl for cl in found_labels)
        assert any("Col B" in cl for cl in found_labels)


# ---------------------------------------------------------------------------
# Footnote filtering
# ---------------------------------------------------------------------------

class TestFootnoteRows:
    def test_prefix_footnotes_stripped(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        ws.cell(row=1, column=1, value="")
        ws.cell(row=1, column=2, value="Total")
        ws.cell(row=1, column=3, value="Male")

        ws.cell(row=2, column=1, value="Aided")
        ws.cell(row=2, column=2, value=80)
        ws.cell(row=2, column=3, value=75)

        ws.cell(row=3, column=1, value="Unaided")
        ws.cell(row=3, column=2, value=40)
        ws.cell(row=3, column=3, value=38)

        ws.cell(row=4, column=1, value="Source: Company Survey 2024")
        ws.cell(row=4, column=2, value=None)
        ws.cell(row=4, column=3, value=None)

        ws.cell(row=5, column=1, value="Note: Data is approximate")
        ws.cell(row=5, column=2, value=None)
        ws.cell(row=5, column=3, value=None)

        path = _save_wb(wb, tmp_path)
        t = parse_workbook(path)["tables"][0]
        for label in t["row_labels"]:
            assert not label.lower().startswith("source:")
            assert not label.lower().startswith("note:")

    def test_substring_footnotes_stripped(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Banner row (dense)
        ws.cell(row=1, column=1, value="")
        ws.cell(row=1, column=2, value="Total")
        ws.cell(row=1, column=3, value="Male")
        ws.cell(row=1, column=4, value="Female")

        # Body rows with data
        ws.cell(row=2, column=1, value="Aided")
        ws.cell(row=2, column=2, value=80)
        ws.cell(row=2, column=3, value=75)
        ws.cell(row=2, column=4, value=85)

        ws.cell(row=3, column=1, value="Unaided")
        ws.cell(row=3, column=2, value=40)
        ws.cell(row=3, column=3, value=38)
        ws.cell(row=3, column=4, value=42)

        # Footnote rows (sparse data — only label, no numeric data)
        ws.cell(row=4, column=1, value="Results are unweighted")

        ws.cell(row=5, column=1, value="Statistical significance at 95%")

        path = _save_wb(wb, tmp_path)
        t = parse_workbook(path)["tables"][0]
        for label in t["row_labels"]:
            assert "unweighted" not in label.lower()
            assert "significance" not in label.lower()


# ---------------------------------------------------------------------------
# Small table near min block threshold
# ---------------------------------------------------------------------------

class TestMinBlockSize:
    def test_small_block_parsed(self, tmp_path):
        """A small but valid block meets min_non_null_cells=4 and 2x2 min dims."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        ws.cell(row=1, column=1, value="")
        ws.cell(row=1, column=2, value="Col A")
        ws.cell(row=1, column=3, value="Col B")
        ws.cell(row=2, column=1, value="Row 1")
        ws.cell(row=2, column=2, value=10)
        ws.cell(row=2, column=3, value=20)
        ws.cell(row=3, column=1, value="Row 2")
        ws.cell(row=3, column=2, value=30)
        ws.cell(row=3, column=3, value=40)

        path = _save_wb(wb, tmp_path)
        result = parse_workbook(path)
        assert len(result["tables"]) >= 1

    def test_1x2_block_skipped(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        ws.cell(row=1, column=1, value="Only one row")
        ws.cell(row=1, column=2, value=42)

        path = _save_wb(wb, tmp_path)
        result = parse_workbook(path, parse_options={"min_block_rows": 2})
        assert len(result["tables"]) == 0


# ---------------------------------------------------------------------------
# Title-in-first-row pattern
# ---------------------------------------------------------------------------

class TestTitleInFirstRow:
    def test_first_row_single_cell_becomes_title(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        ws.cell(row=1, column=1, value="My Custom Title")
        # Row 2: header
        ws.cell(row=2, column=1, value="")
        ws.cell(row=2, column=2, value="Total")
        ws.cell(row=2, column=3, value="Male")

        ws.cell(row=3, column=1, value="Row A")
        ws.cell(row=3, column=2, value=10)
        ws.cell(row=3, column=3, value=20)

        ws.cell(row=4, column=1, value="Row B")
        ws.cell(row=4, column=2, value=30)
        ws.cell(row=4, column=3, value=40)

        path = _save_wb(wb, tmp_path)
        t = parse_workbook(path)["tables"][0]
        assert t["title"] == "My Custom Title"


# ---------------------------------------------------------------------------
# Empty sheet
# ---------------------------------------------------------------------------

class TestEmptySheet:
    def test_empty_sheet_yields_no_tables(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Empty"
        path = _save_wb(wb, tmp_path)
        result = parse_workbook(path)
        assert result["tables"] == []


# ---------------------------------------------------------------------------
# Custom parse_options
# ---------------------------------------------------------------------------

class TestCustomParseOptions:
    def test_high_min_non_null_cells_skips_small_blocks(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        ws.cell(row=1, column=1, value="")
        ws.cell(row=1, column=2, value="Col")
        ws.cell(row=2, column=1, value="Row")
        ws.cell(row=2, column=2, value=42)

        path = _save_wb(wb, tmp_path)
        result = parse_workbook(path, parse_options={"min_non_null_cells": 100})
        assert len(result["tables"]) == 0


# ---------------------------------------------------------------------------
# to_json round-trip
# ---------------------------------------------------------------------------

class TestToJson:
    def test_to_json_roundtrip(self, tmp_path):
        wb = _make_standard_wb()
        path = _save_wb(wb, tmp_path)
        data = parse_workbook(path)
        json_str = to_json(data)
        restored = json.loads(json_str)
        assert restored["tables"][0]["title"] == data["tables"][0]["title"]
        assert len(restored["tables"]) == len(data["tables"])
